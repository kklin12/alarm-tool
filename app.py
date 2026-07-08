# app.py - 完整后端代码（已适配云部署）

# 导入Flask框架相关模块
from flask import Flask, request, jsonify, send_file, render_template_string
import requests
import librosa
import os
import uuid
import matplotlib
import time
import warnings

warnings.filterwarnings('ignore')

matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import sys
import tempfile
import soundfile as sf
import re
import shutil
import scipy.signal

app = Flask(__name__)


# 读取前端HTML页面
def get_resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


# 获取临时目录
def get_temp_dir():
    if getattr(sys, 'frozen', False):
        return tempfile.gettempdir()
    return os.path.abspath(".")


# 临时/输出文件夹初始化
BASE_DIR = get_temp_dir() if getattr(sys, 'frozen', False) else os.path.abspath(".")
OUTPUT = os.path.join(BASE_DIR, "output")
TEMP = os.path.join(BASE_DIR, "temp")
os.makedirs(OUTPUT, exist_ok=True)
os.makedirs(TEMP, exist_ok=True)

# 自己跑可以取消掉注释百度语音密钥
API_KEY = "CpsRtHqvreDhoJIF9gCd94ym"
SECRET_KEY = "xnSfrxfV72Axg2pCH6SysaXPIX8EXOVu"

######上传到云端的
# # 百度语音密钥（云平台环境变量读取，本地.env兼容）
# API_KEY = os.getenv("BAIDU_API_KEY", "")
# SECRET_KEY = os.getenv("BAIDU_SECRET_KEY", "")
# if not API_KEY or not SECRET_KEY:
#     raise RuntimeError("未配置百度语音BAIDU_API_KEY / BAIDU_SECRET_KEY环境变量，请在Render后台填写")
#####


# 音色编号映射（per参数）
PERSON_MAP = {
    0: 0,  # 普通女声（度小美）
    1: 1,  # 普通男声（度小宇）
    2: 3,  # 深沉男声（度逍遥）
    3: 4194,  # 甜美女声（度嫣然）
    4: 4  # 童声（度丫丫）
}
VOICE_NAME_LIST = ["普通女声", "普通男声", "深沉男声", "甜美女声", "童声"]

# 统一音频参数
TARGET_SR = 16000
TARGET_CHANNELS = 1


# ==================== 音频归一化处理 ====================
def normalize_audio(input_path, output_path, target_sr=TARGET_SR, normalize_volume=True):
    """
    将音频统一为：16000Hz、单声道、WAV格式
    解决外部音频格式不一致导致的合并问题
    normalize_volume: False时保留原始音量（用于TTS生成的音频）
    """
    try:
        # 使用 librosa 加载音频（自动处理多种格式）
        y, sr = librosa.load(input_path, sr=None, mono=True)

        # 如果采样率不是目标值，进行重采样
        if sr != target_sr:
            y = librosa.resample(y, orig_sr=sr, target_sr=target_sr)
            sr = target_sr

        # 只在需要时归一化音量
        if normalize_volume:
            max_val = np.max(np.abs(y))
            if max_val > 0:
                y = y / max_val * 0.9

        # 保存为WAV格式
        sf.write(output_path, y, sr, subtype='PCM_16')
        return True
    except Exception as e:
        print(f"音频归一化失败 {input_path}: {e}")
        return False


def normalize_audio_inplace(file_path, target_sr=TARGET_SR):
    """
    原地归一化音频文件
    """
    temp_path = file_path + ".temp.wav"
    if normalize_audio(file_path, temp_path, target_sr):
        # 替换原文件
        if os.path.exists(file_path):
            os.remove(file_path)
        os.rename(temp_path, file_path)
        return True
    return False


# 获取Access Token
def get_token():
    try:
        url = "https://aip.baidubce.com/oauth/2.0/token"
        params = {
            "grant_type": "client_credentials",
            "client_id": API_KEY,
            "client_secret": SECRET_KEY
        }
        res = requests.post(url, params=params, timeout=10)
        return res.json()["access_token"]
    except Exception as e:
        print("Token获取失败：", e)
        return None


# 生成静音wav文件
def create_silence_wav(save_path, duration_sec, sample_rate=TARGET_SR):
    n_samples = int(duration_sec * sample_rate)
    silence = np.zeros(n_samples, dtype=np.float32)
    sf.write(save_path, silence, sample_rate, subtype='PCM_16')
    return save_path


# 拼接多段wav音频（支持不同采样率）
def concat_wav_files(out_path, file_list, target_sr=TARGET_SR):
    all_audio = []
    sr = target_sr

    for fp in file_list:
        if not os.path.exists(fp):
            continue
        # 使用 librosa 加载，统一采样率
        y, s = librosa.load(fp, sr=target_sr, mono=True)
        sr = s
        all_audio.append(y)

    if not all_audio:
        raise ValueError("没有有效的音频文件可拼接")

    full_audio = np.concatenate(all_audio)

    # 归一化防止削波
    max_val = np.max(np.abs(full_audio))
    if max_val > 1.0:
        full_audio = full_audio / max_val * 0.95

    sf.write(out_path, full_audio, sr, subtype='PCM_16')
    return out_path


# 文本预处理：解析 ~ 实现尾音拉长
def process_tilde_long_tone(raw_text):
    pattern = re.compile(r'([\u4e00-\u9fa5])(~+)')

    def replace_func(match):
        char = match.group(1)
        tilde_num = len(match.group(2))
        repeat_times = tilde_num
        return char * repeat_times + "……"

    processed = pattern.sub(replace_func, raw_text)
    return processed


# ==================== 从文件名提取告警文本（改进版） ====================
def extract_alert_text_from_filename(filename):
    """
    从文件名中提取告警文本
    规则：
    1. 去掉文件扩展名
    2. 按分隔符拆分，取第一个非纯数字的片段
    3. 只保留中文和英文字母
    """
    # 去掉扩展名
    name = os.path.splitext(filename)[0]

    # 定义分隔符
    separators = ['_', '-', '—', '，', ',', ' ', '（', '）', '(', ')', '[', ']', '【', '】']

    # 尝试按分隔符拆分
    parts = [name]
    for sep in separators:
        new_parts = []
        for p in parts:
            new_parts.extend(p.split(sep))
        parts = new_parts
        # 如果拆分后已经有多个部分，可以提前结束
        if len(parts) > 1:
            break

    # 过滤掉空字符串
    parts = [p.strip() for p in parts if p.strip()]

    # 从前往后找第一个不是纯数字的片段
    for part in parts:
        # 检查是否是纯数字（包括带小数点的数字）
        if re.match(r'^[\d.]+$', part):
            continue  # 纯数字跳过
        # 检查是否包含中文字符或英文字母
        if re.search(r'[\u4e00-\u9fa5a-zA-Z]', part):
            # 只保留中英文，去掉其他特殊字符（保留空格）
            clean_text = re.sub(r'[^\u4e00-\u9fa5a-zA-Z\s]', '', part)
            clean_text = clean_text.strip()
            if clean_text:  # 如果清理后有内容
                return clean_text

    # 如果所有片段都被过滤掉了，返回原始文件名（去除扩展名）
    # 但也要清理特殊字符
    clean_name = re.sub(r'[^\u4e00-\u9fa5a-zA-Z\s]', '', name)
    clean_name = clean_name.strip()
    return clean_name if clean_name else name


# ==================== 频谱分析辅助函数 ====================
def extract_audio_features(file_path, filename=None):
    """提取音频的完整特征 - 使用pyin算法检测基频"""
    y, sr = librosa.load(file_path, sr=None)

    total_duration = float(librosa.get_duration(y=y, sr=sr))

    # 计算RMS能量
    rms = librosa.feature.rms(y=y)[0]
    rms_mean = float(np.mean(rms))

    # 计算响度（使用RMS的dB值，更准确）
    if rms_mean > 0:
        loudness_db = round(20 * np.log10(rms_mean + 1e-8), 2)
    else:
        loudness_db = -100.0

    # 频谱特征
    spectral_centroid = float(np.mean(librosa.feature.spectral_centroid(y=y, sr=sr)[0]))
    spectral_bandwidth = float(np.mean(librosa.feature.spectral_bandwidth(y=y, sr=sr)[0]))

    # ========== 使用 pyin 计算基频（更准确） ==========
    try:
        # pyin 返回 (f0, voiced_flag, voiced_probs)
        f0, voiced_flag, _ = librosa.pyin(
            y,
            fmin=librosa.note_to_hz('C2'),  # 65.4 Hz
            fmax=librosa.note_to_hz('C6'),  # 1046.5 Hz
            sr=sr
        )
        # 只取有声音的部分
        f0_voiced = f0[voiced_flag]
        if len(f0_voiced) > 0:
            avg_pitch = float(round(np.mean(f0_voiced), 2))
        else:
            avg_pitch = 0.0
    except Exception as e:
        print(f"pyin 计算失败，使用 piptrack 降级: {e}")
        # 降级方案：使用 piptrack 并过滤
        pitches, magnitudes = librosa.piptrack(
            y=y,
            sr=sr,
            fmin=50,
            fmax=500
        )
        pitch_values = pitches[pitches > 0]
        if len(pitch_values) > 0:
            # 只保留人声频率范围 50-500Hz
            pitch_filtered = pitch_values[(pitch_values >= 50) & (pitch_values <= 500)]
            if len(pitch_filtered) > 0:
                avg_pitch = float(round(np.mean(pitch_filtered), 2))
            else:
                avg_pitch = float(round(np.mean(pitch_values), 2))
        else:
            avg_pitch = 0.0

    # 静音检测
    silence_threshold = 0.01
    is_sil = rms < silence_threshold
    total_frames = len(is_sil)

    silence_frames = np.sum(is_sil)
    silence_duration = silence_frames / total_frames * total_duration if total_frames > 0 else 0
    speech_duration = total_duration - silence_duration

    # 告警有效时长 = 语音时长
    alert_effective_duration = speech_duration

    # 语音段比例和静音段比例
    speech_ratio = (speech_duration / total_duration * 100) if total_duration > 0 else 0
    silence_ratio = (silence_duration / total_duration * 100) if total_duration > 0 else 0

    # 采样数和采样频率
    sample_count = len(y)
    sample_rate = sr

    # 自动识别重复告警段
    seg_list = []
    if len(is_sil) > 0:
        current_state = "sil" if is_sil[0] else "voice"
        seg_start = 0
        for idx, s in enumerate(is_sil):
            state = "sil" if s else "voice"
            if state != current_state:
                seg_dur = (idx - seg_start) / total_frames * total_duration if total_frames > 0 else 0
                seg_list.append({"type": current_state, "dur": seg_dur})
                current_state = state
                seg_start = idx
        seg_dur = (total_frames - seg_start) / total_frames * total_duration if total_frames > 0 else 0
        seg_list.append({"type": current_state, "dur": seg_dur})

    voice_segments = [seg["dur"] for seg in seg_list if seg["type"] == "voice"]
    auto_repeat_cnt = len(voice_segments)
    sil_between = []
    for i, seg in enumerate(seg_list):
        if seg["type"] == "sil" and 0 < i < len(seg_list) - 1:
            sil_between.append(seg["dur"])
    auto_repeat_gap = sum(sil_between) / len(sil_between) if sil_between else 0.0

    # 从文件名提取告警文本（使用改进版函数）
    alert_text = extract_alert_text_from_filename(filename) if filename else ""

    return {
        "total_duration": round(total_duration, 2),
        "speech_duration": round(speech_duration, 2),
        "silence_duration": round(silence_duration, 2),
        "alert_effective_duration": round(alert_effective_duration, 2),
        "speech_ratio": round(speech_ratio, 2),
        "silence_ratio": round(silence_ratio, 2),
        "sample_count": sample_count,
        "sample_rate": sample_rate,
        "loudness_db": loudness_db,
        "spectral_centroid": round(spectral_centroid, 2),
        "spectral_bandwidth": round(spectral_bandwidth, 2),
        "avg_pitch": avg_pitch,
        "auto_repeat_cnt": auto_repeat_cnt,
        "auto_repeat_gap_sec": round(auto_repeat_gap, 2),
        "alert_text": alert_text,
        "y": y,
        "sr": sr
    }


# 生成频谱图
def generate_spectrogram(y, sr, filename, output_path):

    ##############################之前的
    plt.rcParams['font.sans-serif'] = ["SimHei", "Microsoft YaHei", "DejaVu Sans"]



    fig, ax = plt.subplots(figsize=(10, 3.5))
    mel_spec = librosa.feature.melspectrogram(y=y, sr=sr, n_mels=128, fmax=8000)
    mel_db = librosa.power_to_db(mel_spec, ref=np.max)
    img = librosa.display.specshow(mel_db, sr=sr, fmax=8000, ax=ax, x_axis="time", y_axis="hz")
    fig.colorbar(img, ax=ax, format="%+2.0f dB", label="相对响度分贝")
    ax.set_title(f"梅尔频谱图：{filename}", fontsize=10, pad=8)
    ax.set_xlabel("时间 (秒 s)", fontsize=9, labelpad=6)
    ax.set_ylabel("梅尔频率 (Hz)", fontsize=9, labelpad=6)
    fig.text(
        0.06, 0.02,
        "亮度越高音量越大，可区分静音与人声段",
        fontsize=8, color="#666666", wrap=True
    )
    plt.tight_layout(pad=1.5, rect=[0, 0.05, 1, 0.97])
    plt.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close(fig)


# ==================== 路由 ====================

@app.route("/")
def index():
    try:
        with open(get_resource_path("index.html"), "r", encoding="utf-8") as f:
            return render_template_string(f.read())
    except Exception as e:
        return f"未找到index.html文件: {str(e)}", 404


@app.route("/api/upload_external", methods=["POST"])
def upload_external():
    """上传外部音频文件到服务器，并归一化格式"""
    try:
        files = request.files.getlist("files")
        results = []

        for f in files:
            # 生成唯一文件名，保留原始扩展名
            ext = os.path.splitext(f.filename)[1]
            if not ext:
                ext = '.wav'
            unique_name = f"external_{uuid.uuid4().hex[:8]}_{int(time.time())}{ext}"

            # 先保存到临时位置
            temp_path = os.path.join(TEMP, f"temp_{uuid.uuid4().hex[:8]}{ext}")
            f.save(temp_path)

            # 最终保存路径（统一为WAV格式）
            final_name = f"external_{uuid.uuid4().hex[:8]}_{int(time.time())}.wav"
            final_path = os.path.join(OUTPUT, final_name)

            # 归一化音频格式（外部音频需要归一化音量）
            if normalize_audio(temp_path, final_path, normalize_volume=True):
                results.append({
                    "original_name": f.filename,
                    "saved_name": final_name,
                    "url": f"/static_output/{final_name}"
                })
            else:
                # 如果归一化失败，尝试直接保存
                shutil.copy2(temp_path, final_path)
                results.append({
                    "original_name": f.filename,
                    "saved_name": final_name,
                    "url": f"/static_output/{final_name}"
                })

            # 清理临时文件
            if os.path.exists(temp_path):
                os.remove(temp_path)

        return jsonify({"success": True, "files": results})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/tts", methods=["POST"])
def tts():
    try:
        data = request.json
        raw_text = data.get("text", "").strip()
        if not raw_text:
            return jsonify({"error": "请输入告警文本"}), 400

        synth_text = process_tilde_long_tone(raw_text)

        voice_gender = int(data.get("voice_gender", 0))
        word_per_min = int(data.get("word_per_min", 240))
        pit_value = int(data.get("pit_value", 5))
        segment_pause = float(data.get("segment_pause", 0.3))
        volume_level = int(data.get("volume_level", 5))
        repeat_times = int(data.get("repeat_times", 1))
        repeat_gap = float(data.get("repeat_gap", 0.8))

        # 参数范围校验
        if not (0 <= voice_gender <= 4):
            return jsonify({"error": "人声类型取值范围0~4"}), 400
        if not (60 <= word_per_min <= 360):
            return jsonify({"error": "每分钟朗读字数范围60~360"}), 400
        if not (0 <= pit_value <= 9):
            return jsonify({"error": "音调等级范围0~9"}), 400
        if not (0 <= segment_pause <= 3):
            return jsonify({"error": "分句停顿时长范围0~3秒"}), 400
        if not (0 <= volume_level <= 15):
            return jsonify({"error": "音量等级范围0~15"}), 400
        if not (1 <= repeat_times <= 10):
            return jsonify({"error": "告警重复次数范围1~10次"}), 400
        if not (0 <= repeat_gap <= 5):
            return jsonify({"error": "告警重复间隔范围0~5秒"}), 400

        # 计算百度TTS语速参数
        base_spd = 5
        word_rate = word_per_min / 240
        spd_value = base_spd * word_rate
        spd_value = max(0, min(15, spd_value))

        # 音量等级直接使用 0-15
        vol = volume_level
        per = PERSON_MAP[voice_gender]
        voice_name = VOICE_NAME_LIST[voice_gender]

        token = get_token()
        if not token:
            return jsonify({"error": "百度语音授权失败，请检查API密钥"}), 500

        # 分句切割
        split_texts = [t.strip() for t in synth_text.replace("。", "，").split("，") if t.strip()]
        seg_audio_paths = []
        seg_sil_path = os.path.join(TEMP, f"seg_sil_{uuid.uuid4()}.wav")
        create_silence_wav(seg_sil_path, segment_pause, TARGET_SR)

        # 逐段调用TTS接口
        for seg_txt in split_texts:
            tts_url = "http://tsn.baidu.com/text2audio"
            params = {
                "tex": seg_txt,
                "lan": "zh",
                "ctp": 1,
                "cuid": "alarm_tts_tool",
                "spd": spd_value,
                "pit": pit_value,
                "vol": vol,
                "per": per,
                "tok": token,
                "aue": 6,
            }

            resp = requests.get(tts_url, params=params, timeout=15)

            try:
                err_info = resp.json()
                return jsonify({"error": f"百度TTS接口报错：{err_info.get('err_msg', '无错误信息')}"}), 400
            except ValueError:
                audio_bin = resp.content

            seg_path = os.path.join(TEMP, f"seg_{uuid.uuid4()}.wav")
            with open(seg_path, "wb") as f:
                f.write(audio_bin)

            # 归一化TTS生成的音频（但保留音量差异，normalize_volume=False）
            norm_path = os.path.join(TEMP, f"seg_norm_{uuid.uuid4()}.wav")
            normalize_audio(seg_path, norm_path, TARGET_SR, normalize_volume=False)
            if os.path.exists(seg_path):
                os.remove(seg_path)

            seg_audio_paths.append(norm_path)
            seg_audio_paths.append(seg_sil_path)

        # 拼接单次完整告警音频
        single_tmp = os.path.join(TEMP, f"single_{uuid.uuid4()}.wav")
        try:
            concat_wav_files(single_tmp, seg_audio_paths[:-1], TARGET_SR)
        except Exception as e:
            return jsonify({"error": f"音频拼接失败：{str(e)}"}), 500

        # 循环重复拼接多次告警
        repeat_sil = os.path.join(TEMP, f"rep_sil_{uuid.uuid4()}.wav")
        create_silence_wav(repeat_sil, repeat_gap, TARGET_SR)
        final_parts = []
        for _ in range(repeat_times):
            final_parts.append(single_tmp)
            final_parts.append(repeat_sil)
        final_parts = final_parts[:-1]

        base_name = re.sub(r'[\\/*?:"<>|]', '', raw_text)
        if len(base_name) > 50:
            base_name = base_name[:50]
        if os.path.exists(os.path.join(OUTPUT, f"{base_name}.wav")):
            base_name = f"{base_name}_{int(time.time())}"
        fname = f"{base_name}.wav"

        full_audio_path = os.path.join(OUTPUT, fname)
        concat_wav_files(full_audio_path, final_parts, TARGET_SR)

        # 清理临时音频文件
        clean_list = seg_audio_paths + [single_tmp, seg_sil_path, repeat_sil]
        for fp in clean_list:
            if os.path.exists(fp):
                try:
                    os.remove(fp)
                except:
                    pass

        return jsonify({
            "file_name": fname,
            "file_url": f"/static_output/{fname}",
            "cfg": {
                "voice_type": voice_name,
                "voice_code": voice_gender,
                "word_per_min": word_per_min,
                "pit_value": pit_value,
                "pause_sec": segment_pause,
                "volume_level": volume_level,
                "repeat_cnt": repeat_times,
                "repeat_gap_sec": repeat_gap,
                "original_text": raw_text,
                "processed_text": synth_text,
                "spd_value": round(spd_value, 2)
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/merge", methods=["POST"])
def merge_audio():
    try:
        data = request.json
        file_names = data.get("file_list", [])
        merge_gap = float(data.get("merge_gap", 0.5))
        if len(file_names) < 2:
            return jsonify({"error": "至少选择2条音频文件进行合并"}), 400
        if not (0 <= merge_gap <= 10):
            return jsonify({"error": "合并中间静默间隔范围0~10秒"}), 400

        file_paths = [os.path.join(OUTPUT, fn) for fn in file_names]
        for fp in file_paths:
            if not os.path.exists(fp):
                return jsonify({"error": f"文件不存在：{os.path.basename(fp)}"}), 400

        # 统一格式化所有待合并音频（保留原始音量）
        normalized_paths = []
        for fp in file_paths:
            # 创建归一化副本
            norm_name = f"norm_{uuid.uuid4().hex[:8]}_{os.path.basename(fp)}"
            norm_path = os.path.join(TEMP, norm_name)

            # 合并时保留原始音量
            if normalize_audio(fp, norm_path, TARGET_SR, normalize_volume=True):
                normalized_paths.append(norm_path)
            else:
                # 如果归一化失败，使用原文件
                normalized_paths.append(fp)

        # 创建合并间隔静音
        sil_path = os.path.join(TEMP, f"merge_sil_{uuid.uuid4()}.wav")
        create_silence_wav(sil_path, merge_gap, TARGET_SR)

        # 构建合并列表
        full_parts = []
        for i, np_path in enumerate(normalized_paths):
            full_parts.append(np_path)
            if i < len(normalized_paths) - 1:
                full_parts.append(sil_path)

        # 合并音频
        merge_name = f"merge_{uuid.uuid4()}.wav"
        merge_path = os.path.join(OUTPUT, merge_name)
        concat_wav_files(merge_path, full_parts, TARGET_SR)

        # 清理临时归一化文件
        for np_path in normalized_paths:
            if np_path != fp and os.path.exists(np_path):
                try:
                    os.remove(np_path)
                except:
                    pass

        if os.path.exists(sil_path):
            os.remove(sil_path)

        return jsonify({
            "merge_file": merge_name,
            "merge_url": f"/static_output/{merge_name}"
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/analyze", methods=["POST"])
def analyze():
    try:
        files = request.files.getlist("files")
        results = []

        for f in files:
            temp_path = os.path.join(TEMP, f"{uuid.uuid4()}.wav")
            f.save(temp_path)

            features = extract_audio_features(temp_path, f.filename)

            base_name = os.path.splitext(f.filename)[0]
            spec_filename = f"spec_{base_name}_{uuid.uuid4().hex[:8]}.png"
            spec_path = os.path.join(OUTPUT, spec_filename)

            generate_spectrogram(features["y"], features["sr"], f.filename, spec_path)

            result = {
                "file": f.filename,
                "alert_text": features["alert_text"],
                "total_duration": features["total_duration"],
                "speech_duration": features["speech_duration"],
                "silence_duration": features["silence_duration"],
                "alert_effective_duration": features["alert_effective_duration"],
                "speech_ratio": features["speech_ratio"],
                "silence_ratio": features["silence_ratio"],
                "sample_count": features["sample_count"],
                "sample_rate": features["sample_rate"],
                "loudness_db": features["loudness_db"],
                "spectral_centroid": features["spectral_centroid"],
                "spectral_bandwidth": features["spectral_bandwidth"],
                "avg_pitch": features["avg_pitch"],
                "auto_repeat_cnt": features["auto_repeat_cnt"],
                "auto_repeat_gap_sec": features["auto_repeat_gap_sec"],
                "spectrogram": f"/static_output/{spec_filename}"
            }
            results.append(result)
            os.remove(temp_path)

        return jsonify({"analysis": results})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/analyze_single", methods=["POST"])
def analyze_single():
    try:
        data = request.json
        filename = data.get("filename", "")
        if not filename:
            return jsonify({"error": "请提供文件名"}), 400

        file_path = os.path.join(OUTPUT, filename)
        if not os.path.exists(file_path):
            return jsonify({"error": "音频文件不存在"}), 404

        temp_path = os.path.join(TEMP, f"{uuid.uuid4()}.wav")
        shutil.copy2(file_path, temp_path)

        features = extract_audio_features(temp_path, filename)

        base_name = os.path.splitext(filename)[0]
        spec_filename = f"spec_{base_name}_{uuid.uuid4().hex[:8]}.png"
        spec_path = os.path.join(OUTPUT, spec_filename)

        generate_spectrogram(features["y"], features["sr"], filename, spec_path)

        os.remove(temp_path)

        result = {
            "file": filename,
            "alert_text": features["alert_text"],
            "total_duration": features["total_duration"],
            "speech_duration": features["speech_duration"],
            "silence_duration": features["silence_duration"],
            "alert_effective_duration": features["alert_effective_duration"],
            "speech_ratio": features["speech_ratio"],
            "silence_ratio": features["silence_ratio"],
            "sample_count": features["sample_count"],
            "sample_rate": features["sample_rate"],
            "loudness_db": features["loudness_db"],
            "spectral_centroid": features["spectral_centroid"],
            "spectral_bandwidth": features["spectral_bandwidth"],
            "avg_pitch": features["avg_pitch"],
            "auto_repeat_cnt": features["auto_repeat_cnt"],
            "auto_repeat_gap_sec": features["auto_repeat_gap_sec"],
            "spectrogram": f"/static_output/{spec_filename}"
        }

        return jsonify({"analysis": [result]})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/export", methods=["POST"])
def export():
    try:
        raw_data = request.json
        analysis_data = raw_data.get("data", [])
        selected_columns = raw_data.get("selected_columns", [])

        if not analysis_data:
            return jsonify({"error": "暂无音频分析数据，请先上传音频分析"}), 400

        # ========== 可勾选的列（不包含评估列） ==========
        optional_columns = [
            {"key": "alert_text", "label": "告警文本"},
            {"key": "voice_type", "label": "人声类型"},
            {"key": "total_duration", "label": "总时长(s)"},
            {"key": "speech_duration", "label": "语音时长(s)"},
            {"key": "alert_effective_duration", "label": "告警有效时长(s)"},
            {"key": "speech_ratio", "label": "语音段比例(%)"},
            {"key": "silence_ratio", "label": "静音段比例(%)"},
            {"key": "sample_count", "label": "采样数"},
            {"key": "sample_rate", "label": "采样频率(Hz)"},
            {"key": "repeat_cnt", "label": "告警重复次数"},
            {"key": "repeat_gap", "label": "告警间隔(s)"},
            {"key": "loudness_db", "label": "实测响度(dB)"},
            {"key": "spectral_centroid", "label": "频谱中心(Hz)"},
            {"key": "spectral_bandwidth", "label": "频谱带宽(Hz)"},
            {"key": "avg_pitch", "label": "平均音调(Hz)"},
        ]

        # ========== 固定导出的评估列 ==========
        fixed_prefix_columns = [
            {"key": "ata_chapter", "label": "ATA章节"},
            {"key": "cas_message", "label": "Aural Alert Message in CAS Definition"},
        ]

        fixed_suffix_columns = [
            {"key": "ata_chapter", "label": "ATA章节"},
            {"key": "cas_message", "label": "Aural Alert Message in CAS Definition"},
            {"key": "a1", "label": "A1 音响告警是否使机组感到恼人?"},
            {"key": "a2", "label": "A2 音响告警传达的紧迫程度是否合理"},
            {"key": "p11", "label": "P1.1 音响告警音色是否合适"},
            {"key": "p12", "label": "P1.2 音响告警重复次数是否合适"},
            {"key": "p13", "label": "P1.3 音响告警持续时长是否合适"},
            {"key": "p14", "label": "P1.4 音响告警音调音色是否合适"},
            {"key": "p21", "label": "P2.1 音响告警语音是否清晰易识别"},
            {"key": "p22", "label": "P2.2 音响告警文案表述是否合理"},
            {"key": "p3", "label": "P3 音响告警关闭方式是否合理"},
            {"key": "suggestion",
             "label": "飞行员建议方案（请列出对应的告警名称及意见，如：\"Altitude Horn：建议修改为XXX\"）"},
            {"key": "remark", "label": "备注"}
        ]

        if not selected_columns:
            selected_columns = [col["key"] for col in optional_columns]

        final_columns = []
        final_columns.extend(fixed_prefix_columns)
        for col in optional_columns:
            if col["key"] in selected_columns:
                final_columns.append(col)
        final_columns.extend(fixed_suffix_columns)

        column_labels = [col["label"] for col in final_columns]
        column_keys = [col["key"] for col in final_columns]

        rows = []
        for idx, item in enumerate(analysis_data, 1):
            row = {
                "ID": idx,
                "ata_chapter": item.get("ata_chapter", ""),
                "cas_message": item.get("cas_message", ""),
                "alert_text": item.get("alert_text", ""),
                "voice_type": item.get("voice_type", ""),
                "total_duration": item.get("total_duration", ""),
                "speech_duration": item.get("speech_duration", ""),
                "alert_effective_duration": item.get("alert_effective_duration", ""),
                "speech_ratio": item.get("speech_ratio", ""),
                "silence_ratio": item.get("silence_ratio", ""),
                "sample_count": item.get("sample_count", ""),
                "sample_rate": item.get("sample_rate", ""),
                "repeat_cnt": item.get("repeat_cnt", item.get("auto_repeat_cnt", 1)),
                "repeat_gap": item.get("repeat_gap", item.get("auto_repeat_gap_sec", 0)),
                "loudness_db": item.get("loudness_db", ""),
                "spectral_centroid": item.get("spectral_centroid", ""),
                "spectral_bandwidth": item.get("spectral_bandwidth", ""),
                "avg_pitch": item.get("avg_pitch", ""),
                "a1": item.get("a1", ""),
                "a2": item.get("a2", ""),
                "p11": item.get("p11", ""),
                "p12": item.get("p12", ""),
                "p13": item.get("p13", ""),
                "p14": item.get("p14", ""),
                "p21": item.get("p21", ""),
                "p22": item.get("p22", ""),
                "p3": item.get("p3", ""),
                "suggestion": item.get("suggestion", ""),
                "remark": item.get("remark", "")
            }
            rows.append(row)

        wb = Workbook()
        ws = wb.active
        ws.title = "机组音响告警飞行员评估表"

        header_labels = ["ID"] + column_labels
        for col_idx, label in enumerate(header_labels, 1):
            ws.cell(row=2, column=col_idx).value = label

        for r_idx, row_data in enumerate(rows, 3):
            ws.cell(row=r_idx, column=1).value = r_idx - 2
            for c_idx, key in enumerate(column_keys, 2):
                ws.cell(row=r_idx, column=c_idx).value = row_data.get(key, "")

        gray_fill = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
        thin_line = Side(style="thin", color="FF666666")
        cell_border = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)
        wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for r in (1, 2):
            for c in range(1, len(header_labels) + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = gray_fill
                cell.border = cell_border
                cell.alignment = wrap_center
                cell.font = Font(bold=True)

        for col_idx in range(1, len(header_labels) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 18

        output_buf = BytesIO()
        wb.save(output_buf)
        output_buf.seek(0)

        return send_file(
            output_buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="机组音响飞行员评估表.xlsx"
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/static_output/<filename>")
def static_output(filename):
    full_path = os.path.join(OUTPUT, filename)
    if not os.path.exists(full_path):
        return "音频文件不存在", 404
    return send_file(full_path)


# ==================== 启动入口 ====================
if __name__ == "__main__":
    # 获取端口（云平台会设置 PORT 环境变量）
    port = int(os.environ.get('PORT', 5001))

    print("=============================================")
    print("✈️ 飞机音响告警语音合成工具")
    print(f"服务运行在端口: {port}")
    print("=============================================")

    # 监听所有网络接口，允许外部访问
    app.run(host="0.0.0.0", port=port, debug=False)